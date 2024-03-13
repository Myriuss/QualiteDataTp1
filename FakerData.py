import openpyxl
from datetime import datetime
from faker import Faker

fake = Faker()

def demander_donnee(message, type_donnee):
    while True:
        try:
            if type_donnee == 'date':
                valeur = input(f"{message} (format JJ/MM/AAAA) : ")
                return datetime.strptime(valeur, "%d/%m/%Y").date()
            elif type_donnee == 'int':
                return int(input(f"{message} : "))
            elif type_donnee == 'float':
                return float(input(f"{message} : "))
            elif type_donnee == 'faker':
                return fake.__getattribute__(message)()
            else:
                return input(f"{message} : ")
        except ValueError:
            print(f"Format invalide pour le type de données {type_donnee}. Réessayez.")

# Remplacez ces valeurs par les vôtres
nom_fichier_excel = 'Classeur1.xlsx'
nom_feuille = 'Feuil1'
nombre_personnes = int(input("Combien de personnes souhaitez-vous ajouter à la base de données ? "))

# Chargez le fichier Excel existant ou créez-le s'il n'existe pas encore
try:
    classeur = openpyxl.load_workbook(nom_fichier_excel)
except FileNotFoundError:
    classeur = openpyxl.Workbook()

# Sélectionnez la feuille existante ou créez-en une nouvelle
if nom_feuille in classeur.sheetnames:
    feuille = classeur[nom_feuille]
else:
    feuille = classeur.create_sheet(title=nom_feuille)

# Ajoutez les données du tableau à la feuille Excel
for _ in range(nombre_personnes):
    donnees_personne = [
        fake.random_int(),
        fake.last_name(),
        fake.first_name(),
        fake.date_of_birth(minimum_age=18, maximum_age=90),
        fake.address(),
        fake.city(),
        fake.company(),
        fake.date_this_decade(),
    ]

    feuille.append(donnees_personne)

# Sauvegardez les modifications dans le fichier Excel
classeur.save(nom_fichier_excel)

print(f"Données enregistrées avec succès dans {nom_fichier_excel}, feuille '{nom_feuille}'.")
