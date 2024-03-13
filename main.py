import openpyxl

# Remplacez ces valeurs par les vôtres
nom_fichier_excel = 'Classeur1.xlsx'
nom_feuille = 'Feuil1'
donnees_tableau = [
    ['Permis ID', 'Nom', 'Prénom', 'Date naiss.', 'adresse', 'Canton', 'Assurance', 'Assuré depuis'],
    [5600021, ' De Miller', 'Davic', '01/07/1965', 'ch. de Chne Latran I', 'Genève', 'Zurich', '01/10/1980'],
    [3400093, 'Rubine', 'Josiane', '26/11/1970', 'avenue de l\'Industrie 23', 'VD', 'Generali', '30/01/1989'],
    [2901775, 'De Miller', 'Christine', '22/05/1969', 'ch. de Chne Latran I', 'Genève', 'TSC', '01/03/1994'],
    [10000056, 'Batini', 'Osvaldo', '24/07/1920', 'route de la Liberté 34', 'Schwyz', 'КРТ', '01/09/2005'],
]

# Chargez le fichier Excel existant ou créez-le s'il n'existe pas encore
try:
    classeur = openpyxl.load_workbook(nom_fichier_excel)
except FileNotFoundError:
    classeur = openpyxl.Workbook()

# Sélectionnez la feuille existante ou créez-en une nouvelle
if nom_feuille in classeur.sheetnames:
    feuille = classeur[nom_feuille]
else:
    feuille = classeur.create_sheet(title=nom_feuille)

# Ajoutez les données du tableau à la feuille Excel
for ligne in donnees_tableau:
    feuille.append(ligne)

# Sauvegardez les modifications dans le fichier Excel
classeur.save(nom_fichier_excel)

print(f"Données enregistrées avec succès dans {nom_fichier_excel}, feuille '{nom_feuille}'.")